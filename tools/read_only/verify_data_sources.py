#!/usr/bin/env python
import sqlite3
from pathlib import Path


def count_sqlite(db_path: Path):
    out = {}
    with sqlite3.connect(str(db_path)) as conn:
        cur = conn.cursor()
        for table in ["client", "direct_sale", "booking", "pending_bill", "entry", "payment"]:
            try:
                cur.execute(f"SELECT COUNT(*) FROM {table}")
                out[table] = int(cur.fetchone()[0])
            except Exception:
                pass
    return out


def count_xlsx(xlsx_path: Path):
    try:
        import openpyxl
    except Exception as exc:
        return {"_error": f"openpyxl not available: {exc}"}

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    out = {}
    for sheet in ["client", "direct_sale", "booking", "pending_bill", "entry", "payment"]:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = 0
        for idx, _ in enumerate(ws.iter_rows(values_only=True), start=1):
            rows = idx
        out[sheet] = max(0, rows - 1)  # subtract header
    return out


def main():
    base = Path(__file__).resolve().parent
    sqlite_path = base / "DATA TO COPY" / "SQLITEBACKUP-06-05-2026_10-55AM.db"
    xlsx_path = base / "DATA TO COPY" / "ALLEXPORT-06-05-2026_10-55AM.xlsx"
    target_sqlite = base / "instance" / "ahmed_cement.db"

    print("DATA SOURCES")
    print("===========")
    if sqlite_path.exists():
        print(f"Source SQLite: {sqlite_path}")
        print(count_sqlite(sqlite_path))
    else:
        print(f"Source SQLite missing: {sqlite_path}")

    if xlsx_path.exists():
        print(f"\nSource XLSX: {xlsx_path}")
        print(count_xlsx(xlsx_path))
    else:
        print(f"\nSource XLSX missing: {xlsx_path}")

    if target_sqlite.exists():
        print(f"\nTarget (app) SQLite: {target_sqlite}")
        print(count_sqlite(target_sqlite))
    else:
        print(f"\nTarget SQLite missing: {target_sqlite}")


if __name__ == "__main__":
    main()
