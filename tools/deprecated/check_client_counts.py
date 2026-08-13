#!/usr/bin/env python
import sqlite3
import openpyxl

source_sqlite = 'DATA TO COPY\\SQLITEBACKUP-06-05-2026_08-40AM.db'
source_xlsx = 'DATA TO COPY\\ALLEXPORT-06-05-2026_09-20AM.xlsx'
target_sqlite = 'instance\\ahmed_cement.db'

print('SOURCE SQLITE CLIENT COUNT:')
with sqlite3.connect(source_sqlite) as conn:
    c = conn.cursor()
    c.execute('SELECT COUNT(*), COUNT(DISTINCT code), COUNT(DISTINCT name) FROM client')
    print(c.fetchone())
    c.execute('SELECT code, COUNT(*) FROM client GROUP BY code HAVING COUNT(*)>1')
    duplicates = c.fetchall()
    print('Duplicate codes in source sqlite:', len(duplicates))
    if duplicates:
        print(duplicates[:10])

print('\nTARGET SQLITE CLIENT COUNT:')
with sqlite3.connect(target_sqlite) as conn:
    c = conn.cursor()
    c.execute('SELECT COUNT(*), COUNT(DISTINCT code), COUNT(DISTINCT name) FROM client')
    print(c.fetchone())
    c.execute('SELECT tenant_id, COUNT(*) FROM client GROUP BY tenant_id')
    print('target tenant counts', c.fetchall())

print('\nSOURCE XLSX CLIENT COUNT:')
wb = openpyxl.load_workbook(source_xlsx, read_only=True)
if 'client' in wb.sheetnames:
    ws = wb['client']
    rows = list(ws.iter_rows(values_only=True))
    print('rows total', max(0, len(rows)-1))
    print('headers', rows[0])
    codes = [r[1] for r in rows[1:] if r and len(r) > 1]
    print('distinct codes', len(set(codes)))
    print('blank code count', sum(1 for r in rows[1:] if not (r and len(r) > 1 and r[1])))
else:
    print('client sheet missing')
