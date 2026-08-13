#!/usr/bin/env python
import openpyxl
import sqlite3

print("=== CHECKING IMPORT FILES ===\n")

# Check XLSX
print("XLSX FILE:")
wb = openpyxl.load_workbook('DATA TO COPY\\ALLEXPORT-06-05-2026_09-20AM.xlsx')
print(f"Sheet Names ({len(wb.sheetnames)} sheets):")
for sheet in wb.sheetnames[:25]:
    ws = wb[sheet]
    print(f"  {sheet}: {ws.dimensions} (rows={ws.max_row}, cols={ws.max_column})")

# Check SQLite
print("\n\nSQLITE DATABASE:")
conn = sqlite3.connect('DATA TO COPY\\SQLITEBACKUP-06-05-2026_08-40AM.db')
cursor = conn.cursor()
cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name;")
tables = [row[0] for row in cursor.fetchall()]
print(f"Tables ({len(tables)}):")
for table in tables[:30]:
    cursor.execute(f"SELECT COUNT(*) FROM {table}")
    count = cursor.fetchone()[0]
    print(f"  {table}: {count} rows")
conn.close()

print("\n=== ANALYSIS COMPLETE ===")
