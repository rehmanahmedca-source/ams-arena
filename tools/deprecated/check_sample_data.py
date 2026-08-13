#!/usr/bin/env python
import openpyxl
import sqlite3

print("=" * 60)
print("SAMPLE DATA FROM XLSX")
print("=" * 60)
wb = openpyxl.load_workbook('DATA TO COPY\\ALLEXPORT-06-05-2026_09-20AM.xlsx')

# Check key tables
for sheet_name in ['client', 'supplier', 'material', 'entry', 'pending_bill']:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{sheet_name.upper()} - Columns & Sample:")
        rows = list(ws.iter_rows(values_only=True, max_row=2))
        if rows:
            print(f"  Headers: {rows[0]}")
            if len(rows) > 1:
                print(f"  Sample:  {rows[1]}")

print("\n\n" + "=" * 60)
print("SAMPLE DATA FROM SQLite")
print("=" * 60)
conn = sqlite3.connect('DATA TO COPY\\SQLITEBACKUP-06-05-2026_08-40AM.db')
cursor = conn.cursor()

# Get schema for key tables
for table in ['client', 'supplier', 'material', 'entry', 'pending_bill']:
    cursor.execute(f"SELECT sql FROM sqlite_master WHERE type='table' AND name='{table}' LIMIT 1")
    result = cursor.fetchone()
    if result:
        print(f"\n{table.upper()} schema:")
        print(f"  {result[0][:120]}...")
        
        cursor.execute(f"SELECT * FROM {table} LIMIT 1")
        cols = [desc[0] for desc in cursor.description]
        row = cursor.fetchone()
        print(f"  Columns: {cols[:8]}")
        if row:
            print(f"  Sample:  {row[:8]}")

conn.close()
